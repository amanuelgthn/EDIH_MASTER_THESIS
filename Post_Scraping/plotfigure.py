import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pycountry
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import plotly.io as pio
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter

# === Step 1: Load Excel data ===
excel_path = 'combined_further_cleaned_keywords.xlsx'
sheet_name = 'Count'
df = pd.read_excel(excel_path, sheet_name=sheet_name)

# === Step 2: Normalize country names and map to ISO-3 ===
def get_iso_alpha3(country_name):
    try:
        return pycountry.countries.lookup(country_name).alpha_3
    except:
        return None

df['ISO'] = df['Country'].apply(get_iso_alpha3)
df = df.dropna(subset=['ISO'])
df['count'] = df['count'].astype(int)

# === Step 3: Create choropleth ===
fig = px.choropleth(
    df,
    locations='ISO',
    color='count',
    hover_name='Country',
    color_continuous_scale=[(0, "red"), (0.5, "yellow"), (1, "green")],
    title='EDIHs Distribution',
)



# === Step 4: Add country count labels via scattergeo ===
# Geolocate country centers
geolocator = Nominatim(user_agent="geoapi")
geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1)

coordinates = []
for country in df['Country']:
    location = geocode(country)
    if location:
        coordinates.append((country, location.latitude, location.longitude))
    else:
        coordinates.append((country, None, None))

coords_df = pd.DataFrame(coordinates, columns=['Country', 'lat', 'lon'])
df_labels = df.merge(coords_df, on='Country').dropna(subset=['lat', 'lon'])

fig.add_trace(go.Scattergeo(
    lon=df_labels['lon'],
    lat=df_labels['lat'],
    text=df_labels['count'],
    mode='text',
    textfont=dict(size=9, color='black'),
    showlegend=False
))

fig.update_geos(
    projection_type="natural earth",
    lataxis_range=[30, 75],
    lonaxis_range=[-25, 60]
)
fig.update_layout(margin={"r":0,"t":40,"l":0,"b":0})

# === Step 5: Save to image in memory ===
img_bytes = pio.to_image(fig, format='png', width=1000, height=600)
img = Image.open(io.BytesIO(img_bytes))
img_path = "temp_map.png"
img.save(img_path)

# # === Step 6: Embed image into Excel file ===
# book = load_workbook(excel_path)
# if "figshow" in book.sheetnames:
#     del book["figshow"]
# ws = book.create_sheet("figshow")

# img_excel = ExcelImage(img_path)
# ws.add_image(img_excel, "A1")
# book.save(excel_path)
